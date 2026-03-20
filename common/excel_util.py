import openpyxl
from config.settings import DATA_DIR
from core.logger import log

class ExcelUtil:
    """Excel文件读写工具（仅支持.xlsx）"""
    def __init__(self, file_name, sheet_name=None):
        self.file_path = DATA_DIR / file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self._load_excel()

    def _load_excel(self):
        """加载Excel文件"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            # 选择sheet：指定名称 / 第一个sheet
            self.sheet = self.workbook[self.sheet_name] if self.sheet_name else self.workbook.active
            log.info(f"加载Excel文件成功：{self.file_path}，sheet={self.sheet.title}")
        except Exception as e:
            log.error(f"加载Excel文件失败：{str(e)}")
            raise

    def read_all_data(self):
        """读取所有数据（返回列表套字典）"""
        # 获取表头
        headers = [cell.value for cell in self.sheet[1]]
        # 读取数据行
        data_list = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data_list.append(row_dict)
        log.info(f"读取Excel数据条数：{len(data_list)}")
        return data_list

    def write_cell(self, row, col, value):
        """写入单元格"""
        try:
            self.sheet.cell(row=row, column=col, value=value)
            self.workbook.save(self.file_path)
            log.info(f"写入Excel成功：行{row}列{col} = {value}")
        except Exception as e:
            log.error(f"写入Excel失败：{str(e)}")
            raise

    def close(self):
        """关闭Excel文件"""
        if self.workbook:
            self.workbook.close()

# 示例用法
if __name__ == "__main__":
    excel = ExcelUtil("test_data.xlsx", "login")
    print(excel.read_all_data())
    excel.close()